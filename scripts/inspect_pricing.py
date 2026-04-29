"""Inspect the price book to extract the direct ARR inputs formula."""
import openpyxl
from pathlib import Path

XLSX = Path(r"C:/Users/MikeGuadan/Downloads/Public Safety Offering Price Book.xlsx")

wb = openpyxl.load_workbook(XLSX, data_only=False)
print("SHEETS:", wb.sheetnames)
print()

for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== {name} ({ws.max_row} rows × {ws.max_column} cols) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=False):
        cells = [(c.coordinate, c.value) for c in row if c.value is not None]
        if cells:
            print(" | ".join(f"{coord}={val}" for coord, val in cells))
    print()
